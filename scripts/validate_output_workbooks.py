#!/usr/bin/env python3
"""Validate generated workbooks against their BigQuery source tables."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from google.api_core.exceptions import GoogleAPIError, NotFound
from google.cloud import bigquery
from openpyxl import load_workbook

SHEET_TABLES = {
    "sales": "access_input_sales",
    "store_detail": "access_input_store_detail",
    "author_conditions": "access_input_author_conditions",
    "pod_sales": "access_input_pod_sales",
}
EXPECTED_SHEETS = {
    "access": ["sales", "store_detail", "author_conditions"],
    "pod": ["pod_sales"],
}
NUMERIC_BIGQUERY_TYPES = {"INTEGER", "INT64", "FLOAT", "FLOAT64", "NUMERIC", "BIGNUMERIC", "DECIMAL", "BIGDECIMAL"}
ABSOLUTE_TOLERANCE = Decimal("0.000001")
RELATIVE_TOLERANCE = Decimal("0.000000001")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--access", required=True)
    parser.add_argument("--pod", required=True)
    parser.add_argument("--project-id", required=True)
    parser.add_argument("--dataset", required=True)
    parser.add_argument("--location", default="asia-northeast1")
    parser.add_argument("--json-output", required=True)
    parser.add_argument("--markdown-output", required=True)
    return parser.parse_args()


def normalize_header(value: Any) -> str:
    return "" if value is None else str(value).strip()


def decimal_value(value: Any) -> Decimal | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    try:
        result = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError, AttributeError):
        return None
    return result if result.is_finite() else None


def decimal_text(value: Decimal | None) -> str | None:
    if value is None:
        return None
    return format(value, "f")


def decimals_match(left: Decimal, right: Decimal) -> bool:
    difference = abs(left - right)
    scale = max(abs(left), abs(right), Decimal("1"))
    return difference <= max(ABSOLUTE_TOLERANCE, RELATIVE_TOLERANCE * scale)


def inspect_workbook(path: Path, kind: str) -> dict[str, Any]:
    expected = EXPECTED_SHEETS[kind]
    result: dict[str, Any] = {
        "kind": kind,
        "file": path.name,
        "path": str(path),
        "exists": path.exists(),
        "expected_sheets": expected,
        "actual_sheets": [],
        "missing_sheets": [],
        "unexpected_sheets": [],
        "sheets": [],
        "errors": [],
        "warnings": [],
    }
    if not path.exists():
        result["errors"].append("file_not_found")
        return result

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        result["errors"].append(f"workbook_open_failed:{type(exc).__name__}")
        return result

    actual = workbook.sheetnames
    result["actual_sheets"] = actual
    result["missing_sheets"] = [name for name in expected if name not in actual]
    result["unexpected_sheets"] = [name for name in actual if name not in expected]
    if result["missing_sheets"]:
        result["errors"].append("missing_expected_sheets")
    if result["unexpected_sheets"]:
        result["warnings"].append("unexpected_sheets")

    for sheet_name in actual:
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        header_row = next(rows, ())
        headers = [normalize_header(value) for value in header_row]
        duplicate_headers = sorted(name for name, count in Counter(headers).items() if name and count > 1)
        blank_headers = sum(1 for name in headers if not name)
        data_rows = 0
        completely_blank_rows = 0
        null_cells = 0
        non_null_cells = 0

        for row in rows:
            data_rows += 1
            values = list(row)
            if not any(value not in (None, "") for value in values):
                completely_blank_rows += 1
            for value in values[: len(headers)]:
                if value in (None, ""):
                    null_cells += 1
                else:
                    non_null_cells += 1

        sheet_result = {
            "name": sheet_name,
            "table": SHEET_TABLES.get(sheet_name),
            "columns": len(headers),
            "rows": data_rows,
            "headers": headers,
            "duplicate_headers": duplicate_headers,
            "blank_headers": blank_headers,
            "completely_blank_rows": completely_blank_rows,
            "null_cells": null_cells,
            "non_null_cells": non_null_cells,
            "bigquery": None,
            "comparison": None,
            "numeric_comparison": [],
        }
        if not headers:
            result["errors"].append(f"{sheet_name}:missing_header")
        if duplicate_headers:
            result["errors"].append(f"{sheet_name}:duplicate_headers")
        if blank_headers:
            result["errors"].append(f"{sheet_name}:blank_headers")
        if data_rows == 0:
            result["warnings"].append(f"{sheet_name}:zero_data_rows")
        if completely_blank_rows:
            result["warnings"].append(f"{sheet_name}:blank_data_rows")
        if sheet_name not in SHEET_TABLES:
            result["warnings"].append(f"{sheet_name}:no_bigquery_mapping")
        result["sheets"].append(sheet_result)

    workbook.close()
    return result


def workbook_numeric_totals(path: Path, sheet_name: str, numeric_columns: list[str]) -> dict[str, Decimal]:
    totals = {column: Decimal("0") for column in numeric_columns}
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = [normalize_header(value) for value in next(rows, ())]
    positions = {name: index for index, name in enumerate(headers) if name in totals}

    for row in rows:
        for column, index in positions.items():
            value = decimal_value(row[index] if index < len(row) else None)
            if value is not None:
                totals[column] += value

    workbook.close()
    return totals


def bigquery_numeric_totals(
    client: bigquery.Client, table_id: str, numeric_columns: list[str], location: str
) -> dict[str, Decimal]:
    if not numeric_columns:
        return {}
    expressions = [
        f"CAST(COALESCE(SUM(CAST(`{column}` AS BIGNUMERIC)), 0) AS STRING) AS `{column}`"
        for column in numeric_columns
    ]
    sql = f"SELECT {', '.join(expressions)} FROM `{table_id}`"
    row = next(iter(client.query(sql, location=location).result()))
    return {column: Decimal(str(row[column])) for column in numeric_columns}


def attach_bigquery_checks(
    workbooks: list[dict[str, Any]], project_id: str, dataset: str, location: str
) -> list[str]:
    client = bigquery.Client(project=project_id, location=location)
    global_errors: list[str] = []

    for workbook in workbooks:
        path = Path(workbook["path"])
        for sheet in workbook["sheets"]:
            table_name = sheet.get("table")
            if not table_name:
                continue
            table_id = f"{project_id}.{dataset}.{table_name}"
            try:
                table = client.get_table(table_id)
                bq_headers = [field.name for field in table.schema]
                numeric_columns = [
                    field.name for field in table.schema if field.field_type.upper() in NUMERIC_BIGQUERY_TYPES
                ]
                count_query = client.query(f"SELECT COUNT(*) AS row_count FROM `{table_id}`", location=location)
                bq_rows = int(next(iter(count_query.result()))["row_count"])
                bq_totals = bigquery_numeric_totals(client, table_id, numeric_columns, location)
                excel_totals = workbook_numeric_totals(path, sheet["name"], numeric_columns)
            except NotFound:
                error = f"{sheet['name']}:bigquery_table_not_found"
                workbook["errors"].append(error)
                global_errors.append(error)
                sheet["bigquery"] = {"table_id": table_id, "exists": False}
                continue
            except (GoogleAPIError, InvalidOperation, ValueError) as exc:
                error = f"{sheet['name']}:bigquery_query_failed:{type(exc).__name__}"
                workbook["errors"].append(error)
                global_errors.append(error)
                sheet["bigquery"] = {"table_id": table_id, "exists": None}
                continue

            row_count_match = sheet["rows"] == bq_rows
            column_order_match = sheet["headers"] == bq_headers
            numeric_results = []
            for column in numeric_columns:
                excel_total = excel_totals[column]
                bq_total = bq_totals[column]
                difference = excel_total - bq_total
                matches = decimals_match(excel_total, bq_total)
                numeric_results.append(
                    {
                        "column": column,
                        "workbook_total": decimal_text(excel_total),
                        "bigquery_total": decimal_text(bq_total),
                        "difference": decimal_text(difference),
                        "match": matches,
                    }
                )
                if not matches:
                    error = f"{sheet['name']}:{column}:numeric_total_mismatch"
                    workbook["errors"].append(error)
                    global_errors.append(error)

            sheet["bigquery"] = {
                "table_id": table_id,
                "exists": True,
                "rows": bq_rows,
                "columns": len(bq_headers),
                "headers": bq_headers,
                "numeric_columns": numeric_columns,
            }
            sheet["comparison"] = {
                "row_count_match": row_count_match,
                "column_order_match": column_order_match,
                "numeric_totals_match": all(result["match"] for result in numeric_results),
                "workbook_rows": sheet["rows"],
                "bigquery_rows": bq_rows,
                "workbook_columns": len(sheet["headers"]),
                "bigquery_columns": len(bq_headers),
            }
            sheet["numeric_comparison"] = numeric_results
            if not row_count_match:
                error = f"{sheet['name']}:row_count_mismatch"
                workbook["errors"].append(error)
                global_errors.append(error)
            if not column_order_match:
                error = f"{sheet['name']}:column_order_mismatch"
                workbook["errors"].append(error)
                global_errors.append(error)

    return global_errors


def build_markdown(report: dict[str, Any]) -> str:
    lines = [
        "## Generated workbook validation",
        "",
        f"**Result:** {'OK' if report['ok'] else 'NG'}",
        "",
        "### Workbook and BigQuery comparison",
        "",
        "| Workbook | Sheet | Excel rows | BQ rows | Rows | Excel cols | BQ cols | Column order | Numeric totals |",
        "|---|---|---:|---:|---|---:|---:|---|---|",
    ]
    for workbook in report["workbooks"]:
        if not workbook["sheets"]:
            lines.append(f"| {workbook['file']} | unavailable | 0 | - | NG | 0 | - | NG | NG |")
        for sheet in workbook["sheets"]:
            bq = sheet.get("bigquery") or {}
            comparison = sheet.get("comparison") or {}
            lines.append(
                f"| {workbook['file']} | {sheet['name']} | {sheet['rows']} | {bq.get('rows', '-')} | "
                f"{'OK' if comparison.get('row_count_match') else 'NG'} | {sheet['columns']} | "
                f"{bq.get('columns', '-')} | {'OK' if comparison.get('column_order_match') else 'NG'} | "
                f"{'OK' if comparison.get('numeric_totals_match') else 'NG'} |"
            )

    numeric_rows = [
        (sheet["name"], result)
        for workbook in report["workbooks"]
        for sheet in workbook["sheets"]
        for result in sheet.get("numeric_comparison", [])
    ]
    if numeric_rows:
        lines.extend(
            [
                "",
                "### Numeric totals",
                "",
                "| Sheet | Column | Excel total | BigQuery total | Difference | Result |",
                "|---|---|---:|---:|---:|---|",
            ]
        )
        for sheet_name, result in numeric_rows:
            lines.append(
                f"| {sheet_name} | {result['column']} | {result['workbook_total']} | "
                f"{result['bigquery_total']} | {result['difference']} | {'OK' if result['match'] else 'NG'} |"
            )

    lines.extend(["", "### Structural checks", ""])
    for workbook in report["workbooks"]:
        missing = ", ".join(workbook["missing_sheets"]) or "none"
        unexpected = ", ".join(workbook["unexpected_sheets"]) or "none"
        errors = ", ".join(workbook["errors"]) or "none"
        warnings = ", ".join(workbook["warnings"]) or "none"
        lines.extend(
            [
                f"- **{workbook['file']}**",
                f"  - Missing sheets: {missing}",
                f"  - Unexpected sheets: {unexpected}",
                f"  - Errors: {errors}",
                f"  - Warnings: {warnings}",
            ]
        )
    return "\n".join(lines) + "\n"


def main() -> None:
    args = parse_args()
    workbooks = [inspect_workbook(Path(args.access), "access"), inspect_workbook(Path(args.pod), "pod")]
    bigquery_errors = attach_bigquery_checks(workbooks, args.project_id, args.dataset, args.location)
    report = {
        "ok": all(not workbook["errors"] for workbook in workbooks),
        "project_id": args.project_id,
        "dataset": args.dataset,
        "location": args.location,
        "numeric_tolerance": {
            "absolute": decimal_text(ABSOLUTE_TOLERANCE),
            "relative": decimal_text(RELATIVE_TOLERANCE),
        },
        "error_count": sum(len(workbook["errors"]) for workbook in workbooks),
        "warning_count": sum(len(workbook["warnings"]) for workbook in workbooks),
        "bigquery_error_count": len(bigquery_errors),
        "workbooks": workbooks,
    }

    json_path = Path(args.json_output)
    markdown_path = Path(args.markdown_output)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    markdown_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    markdown_path.write_text(build_markdown(report), encoding="utf-8")
    print(json.dumps({"ok": report["ok"], "errors": report["error_count"], "warnings": report["warning_count"]}))

    if not report["ok"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
