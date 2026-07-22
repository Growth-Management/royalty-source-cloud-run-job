#!/usr/bin/env python3
"""Compare required-null and duplicate-key counts between output workbooks and BigQuery."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from google.cloud import bigquery
from openpyxl import load_workbook

SHEET_CONFIG = {
    "sales": {
        "workbook": "access",
        "table": "access_input_sales",
        "required": [
            "accounting_month_1",
            "accounting_month_2",
            "billing_code_1",
            "product_code",
            "product_key",
            "source_file_id",
            "source_row_number",
        ],
        "key": ["source_file_id", "source_row_number"],
    },
    "store_detail": {
        "workbook": "access",
        "table": "access_input_store_detail",
        "required": [
            "accounting_month_1",
            "accounting_month_2",
            "billing_code",
            "store_name",
            "product_key",
            "source_file_name",
        ],
        "key": ["source_file_name", "source_row_number", "product_key"],
    },
    "author_conditions": {
        "workbook": "access",
        "table": "access_input_author_conditions",
        "required": ["product_key", "product_code", "author_identifier_id", "author_name"],
        "key": ["product_code", "author_identifier_id"],
    },
    "pod_sales": {
        "workbook": "pod",
        "table": "access_input_pod_sales",
        "required": [
            "accounting_month",
            "product_code",
            "product_key",
            "billing_code",
            "source_file_id",
            "source_row_number",
        ],
        "key": ["source_file_id", "source_row_number"],
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--access", required=True)
    parser.add_argument("--pod", required=True)
    parser.add_argument("--project-id", required=True)
    parser.add_argument("--dataset", required=True)
    parser.add_argument("--location", default="asia-northeast1")
    parser.add_argument("--json-output", required=True)
    parser.add_argument("--markdown-output", required=True)
    return parser.parse_args()


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def normalized_key_value(value: Any) -> str:
    if is_blank(value):
        return "<NULL>"
    return str(value).strip()


def workbook_counts(path: Path, sheet_name: str, required: list[str], key: list[str]) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    headers = ["" if value is None else str(value).strip() for value in next(rows, ())]
    positions = {name: index for index, name in enumerate(headers)}
    missing_columns = [column for column in required + key if column not in positions]
    if missing_columns:
        workbook.close()
        return {
            "missing_columns": sorted(set(missing_columns)),
            "required_null_counts": {},
            "duplicate_key_groups": None,
            "duplicate_rows": None,
            "duplicate_samples": [],
        }

    null_counts = {column: 0 for column in required}
    key_counts: Counter[tuple[str, ...]] = Counter()
    for row in rows:
        for column in required:
            index = positions[column]
            value = row[index] if index < len(row) else None
            if is_blank(value):
                null_counts[column] += 1
        key_value = tuple(
            normalized_key_value(row[positions[column]] if positions[column] < len(row) else None)
            for column in key
        )
        key_counts[key_value] += 1

    duplicates = [(values, count) for values, count in key_counts.items() if count > 1]
    workbook.close()
    return {
        "missing_columns": [],
        "required_null_counts": null_counts,
        "duplicate_key_groups": len(duplicates),
        "duplicate_rows": sum(count - 1 for _, count in duplicates),
        "duplicate_samples": [
            {"key": dict(zip(key, values)), "count": count} for values, count in duplicates[:5]
        ],
    }


def bigquery_counts(
    client: bigquery.Client,
    table_id: str,
    required: list[str],
    key: list[str],
    location: str,
) -> dict[str, Any]:
    null_expressions = [
        f"COUNTIF(`{column}` IS NULL OR TRIM(CAST(`{column}` AS STRING)) = '') AS `{column}`"
        for column in required
    ]
    null_row = next(
        iter(client.query(f"SELECT {', '.join(null_expressions)} FROM `{table_id}`", location=location).result())
    )
    null_counts = {column: int(null_row[column]) for column in required}

    key_columns = ", ".join(f"`{column}`" for column in key)
    duplicate_sql = f"""
        WITH duplicate_keys AS (
          SELECT {key_columns}, COUNT(*) AS duplicate_count
          FROM `{table_id}`
          GROUP BY {key_columns}
          HAVING COUNT(*) > 1
        )
        SELECT
          COUNT(*) AS duplicate_key_groups,
          COALESCE(SUM(duplicate_count - 1), 0) AS duplicate_rows,
          TO_JSON_STRING(ARRAY_AGG(STRUCT({key_columns}, duplicate_count) LIMIT 5)) AS duplicate_samples
        FROM duplicate_keys
    """
    duplicate_row = next(iter(client.query(duplicate_sql, location=location).result()))
    samples_json = duplicate_row["duplicate_samples"]
    return {
        "required_null_counts": null_counts,
        "duplicate_key_groups": int(duplicate_row["duplicate_key_groups"]),
        "duplicate_rows": int(duplicate_row["duplicate_rows"]),
        "duplicate_samples": json.loads(samples_json) if samples_json else [],
    }


def build_markdown(report: dict[str, Any]) -> str:
    lines = [
        "## Required values and duplicate keys",
        "",
        f"**Result:** {'OK' if report['ok'] else 'NG'}",
        "",
        "### Required NULL counts",
        "",
        "| Sheet | Column | Excel | BigQuery | Counts match | Required valid |",
        "|---|---|---:|---:|---|---|",
    ]
    for sheet in report["sheets"]:
        for result in sheet["required_results"]:
            lines.append(
                f"| {sheet['sheet']} | {result['column']} | {result['workbook_count']} | "
                f"{result['bigquery_count']} | {'OK' if result['counts_match'] else 'NG'} | "
                f"{'OK' if result['valid'] else 'NG'} |"
            )

    lines.extend(
        [
            "",
            "### Duplicate keys",
            "",
            "| Sheet | Key | Excel groups | BQ groups | Excel duplicate rows | BQ duplicate rows | Counts match | Unique |",
            "|---|---|---:|---:|---:|---:|---|---|",
        ]
    )
    for sheet in report["sheets"]:
        duplicate = sheet["duplicate_result"]
        lines.append(
            f"| {sheet['sheet']} | {', '.join(sheet['key'])} | {duplicate['workbook_groups']} | "
            f"{duplicate['bigquery_groups']} | {duplicate['workbook_rows']} | {duplicate['bigquery_rows']} | "
            f"{'OK' if duplicate['counts_match'] else 'NG'} | {'OK' if duplicate['unique'] else 'NG'} |"
        )

    errors = report.get("errors", [])
    lines.extend(["", "### Errors", "", ", ".join(errors) if errors else "none"])
    return "\n".join(lines) + "\n"


def main() -> None:
    args = parse_args()
    workbook_paths = {"access": Path(args.access), "pod": Path(args.pod)}
    client = bigquery.Client(project=args.project_id, location=args.location)
    sheets = []
    errors: list[str] = []

    for sheet_name, config in SHEET_CONFIG.items():
        workbook_result = workbook_counts(
            workbook_paths[config["workbook"]], sheet_name, config["required"], config["key"]
        )
        if workbook_result["missing_columns"]:
            errors.append(f"{sheet_name}:missing_columns:{','.join(workbook_result['missing_columns'])}")
            sheets.append(
                {
                    "sheet": sheet_name,
                    "table": config["table"],
                    "required": config["required"],
                    "key": config["key"],
                    "required_results": [],
                    "duplicate_result": {
                        "workbook_groups": None,
                        "bigquery_groups": None,
                        "workbook_rows": None,
                        "bigquery_rows": None,
                        "counts_match": False,
                        "unique": False,
                    },
                }
            )
            continue

        table_id = f"{args.project_id}.{args.dataset}.{config['table']}"
        bq_result = bigquery_counts(client, table_id, config["required"], config["key"], args.location)
        required_results = []
        for column in config["required"]:
            workbook_count = workbook_result["required_null_counts"][column]
            bq_count = bq_result["required_null_counts"][column]
            counts_match = workbook_count == bq_count
            valid = workbook_count == 0 and bq_count == 0
            required_results.append(
                {
                    "column": column,
                    "workbook_count": workbook_count,
                    "bigquery_count": bq_count,
                    "counts_match": counts_match,
                    "valid": valid,
                }
            )
            if not counts_match:
                errors.append(f"{sheet_name}:{column}:required_null_count_mismatch")
            if not valid:
                errors.append(f"{sheet_name}:{column}:required_null_present")

        duplicate_counts_match = (
            workbook_result["duplicate_key_groups"] == bq_result["duplicate_key_groups"]
            and workbook_result["duplicate_rows"] == bq_result["duplicate_rows"]
        )
        unique = workbook_result["duplicate_key_groups"] == 0 and bq_result["duplicate_key_groups"] == 0
        if not duplicate_counts_match:
            errors.append(f"{sheet_name}:duplicate_key_count_mismatch")
        if not unique:
            errors.append(f"{sheet_name}:duplicate_key_present")

        sheets.append(
            {
                "sheet": sheet_name,
                "table": table_id,
                "required": config["required"],
                "key": config["key"],
                "required_results": required_results,
                "duplicate_result": {
                    "workbook_groups": workbook_result["duplicate_key_groups"],
                    "bigquery_groups": bq_result["duplicate_key_groups"],
                    "workbook_rows": workbook_result["duplicate_rows"],
                    "bigquery_rows": bq_result["duplicate_rows"],
                    "counts_match": duplicate_counts_match,
                    "unique": unique,
                    "workbook_samples": workbook_result["duplicate_samples"],
                    "bigquery_samples": bq_result["duplicate_samples"],
                },
            }
        )

    report = {
        "ok": not errors,
        "project_id": args.project_id,
        "dataset": args.dataset,
        "errors": errors,
        "error_count": len(errors),
        "sheets": sheets,
    }
    json_path = Path(args.json_output)
    markdown_path = Path(args.markdown_output)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    markdown_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    markdown_path.write_text(build_markdown(report), encoding="utf-8")
    print(json.dumps({"ok": report["ok"], "errors": report["error_count"]}))
    if not report["ok"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
